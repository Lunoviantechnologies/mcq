"""
Excel import/export utilities for Quiz bulk operations
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib.auth.models import User
from .models import Quiz, Question, Choice


def export_quizzes_to_excel(queryset):
    """Export quizzes with questions and choices to Excel format"""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create Quiz sheet
    quiz_sheet = wb.create_sheet("Quizzes")
    quiz_headers = ['Quiz Title', 'Description', 'Created By', 'Is Active', 'Created At']
    quiz_sheet.append(quiz_headers)
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(quiz_headers, 1):
        cell = quiz_sheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add quiz data
    for quiz in queryset:
        quiz_sheet.append([
            quiz.title,
            quiz.description,
            quiz.created_by.username,
            'Yes' if quiz.is_active else 'No',
            quiz.created_at.strftime('%Y-%m-%d %H:%M:%S') if quiz.created_at else ''
        ])
    
    # Auto-adjust column widths for Quiz sheet
    for col in quiz_sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        quiz_sheet.column_dimensions[col_letter].width = adjusted_width
    
    # Create Questions sheet
    question_sheet = wb.create_sheet("Questions")
    question_headers = ['Quiz Title', 'Question Text', 'Question Type', 'Order', 'Time Limit (minutes)']
    question_sheet.append(question_headers)
    
    # Style question headers
    for col_num, header in enumerate(question_headers, 1):
        cell = question_sheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add questions data
    for quiz in queryset:
        for question in quiz.questions.all().order_by('order'):
            question_sheet.append([
                quiz.title,
                question.question_text,
                question.get_question_type_display(),
                question.order,
                question.time_limit_minutes
            ])
    
    # Auto-adjust column widths for Questions sheet
    for col in question_sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 80)
        question_sheet.column_dimensions[col_letter].width = adjusted_width
    
    # Create Choices sheet
    choice_sheet = wb.create_sheet("Choices")
    choice_headers = ['Quiz Title', 'Question Text', 'Choice Text', 'Is Correct']
    choice_sheet.append(choice_headers)
    
    # Style choice headers
    for col_num, header in enumerate(choice_headers, 1):
        cell = choice_sheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add choices data
    for quiz in queryset:
        for question in quiz.questions.filter(question_type='multiple_choice').order_by('order'):
            for choice in question.choices.all():
                choice_sheet.append([
                    quiz.title,
                    question.question_text,
                    choice.choice_text,
                    'Yes' if choice.is_correct else 'No'
                ])
    
    # Auto-adjust column widths for Choices sheet
    for col in choice_sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 80)
        choice_sheet.column_dimensions[col_letter].width = adjusted_width
    
    return wb


def create_excel_template():
    """Create an Excel template for bulk import"""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create Instructions sheet
    instructions_sheet = wb.create_sheet("Instructions", 0)
    instructions = [
        ["QUIZ BULK IMPORT TEMPLATE - INSTRUCTIONS"],
        [""],
        ["HOW TO USE THIS TEMPLATE:"],
        [""],
        ["1. QUIZ SHEET:"],
        ["   - Fill in Quiz Title (required)"],
        ["   - Fill in Description (optional)"],
        ["   - Created By: Enter username (defaults to admin if user not found)"],
        ["   - Is Active: Enter 'Yes' or 'No' (defaults to 'Yes')"],
        [""],
        ["2. QUESTIONS SHEET:"],
        ["   - Quiz Title: Must match exactly with Quiz Title in Quizzes sheet"],
        ["   - Question Text: The question content (required)"],
        ["   - Question Type: 'Multiple Choice' or 'Text Answer' (defaults to 'Multiple Choice')"],
        ["   - Order: Question number/order (0, 1, 2, ...)"],
        ["   - Time Limit: Minutes allowed for question (defaults to 1)"],
        [""],
        ["3. CHOICES SHEET:"],
        ["   - Quiz Title: Must match Quiz Title"],
        ["   - Question Text: Must match Question Text from Questions sheet"],
        ["   - Choice Text: The answer option text"],
        ["   - Is Correct: Enter 'Yes' or 'No' (only for Multiple Choice questions)"],
        [""],
        ["IMPORTANT NOTES:"],
        ["- Multiple Choice questions MUST have at least 2 choices"],
        ["- At least one choice per Multiple Choice question should have 'Is Correct' = 'Yes'"],
        ["- Text Answer questions don't need entries in Choices sheet"],
        ["- Order numbers should be sequential (0, 1, 2, ...)"],
        ["- All Quiz Titles must match exactly across all sheets"],
    ]
    
    for row_idx, instruction in enumerate(instructions, 1):
        cell = instructions_sheet.cell(row=row_idx, column=1, value=instruction[0])
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif instruction[0] and instruction[0].isupper():
            cell.font = Font(bold=True)
    
    instructions_sheet.column_dimensions['A'].width = 100
    
    # Create Quizzes sheet with headers
    quiz_sheet = wb.create_sheet("Quizzes")
    quiz_headers = ['Quiz Title', 'Description', 'Created By', 'Is Active']
    quiz_sheet.append(quiz_headers)
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(quiz_headers, 1):
        cell = quiz_sheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add example row
    quiz_sheet.append([
        'Sample Quiz',
        'This is a sample quiz description',
        'admin',
        'Yes'
    ])
    
    # Auto-adjust column widths
    for col in quiz_sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        quiz_sheet.column_dimensions[col_letter].width = adjusted_width
    
    # Create Questions sheet with headers
    question_sheet = wb.create_sheet("Questions")
    question_headers = ['Quiz Title', 'Question Text', 'Question Type', 'Order', 'Time Limit (minutes)']
    question_sheet.append(question_headers)
    
    # Style headers
    for col_num, header in enumerate(question_headers, 1):
        cell = question_sheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add example rows
    question_sheet.append([
        'Sample Quiz',
        'What is Python?',
        'Multiple Choice',
        '0',
        '1'
    ])
    question_sheet.append([
        'Sample Quiz',
        'Explain Python in your own words.',
        'Text Answer',
        '1',
        '2'
    ])
    
    # Auto-adjust column widths
    for col in question_sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 80)
        question_sheet.column_dimensions[col_letter].width = adjusted_width
    
    # Create Choices sheet with headers
    choice_sheet = wb.create_sheet("Choices")
    choice_headers = ['Quiz Title', 'Question Text', 'Choice Text', 'Is Correct']
    choice_sheet.append(choice_headers)
    
    # Style headers
    for col_num, header in enumerate(choice_headers, 1):
        cell = choice_sheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add example rows
    choice_sheet.append(['Sample Quiz', 'What is Python?', 'A snake', 'No'])
    choice_sheet.append(['Sample Quiz', 'What is Python?', 'A programming language', 'Yes'])
    choice_sheet.append(['Sample Quiz', 'What is Python?', 'A type of pie', 'No'])
    
    # Auto-adjust column widths
    for col in choice_sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 80)
        choice_sheet.column_dimensions[col_letter].width = adjusted_width
    
    return wb


def import_quizzes_from_excel(excel_file, created_by_user):
    """Import quizzes, questions, and choices from Excel file"""
    errors = []
    success_count = 0
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
    except Exception as e:
        return {'success': False, 'errors': [f'Error reading Excel file: {str(e)}'], 'success_count': 0}
    
    # Get default user if created_by_user is not provided
    if not created_by_user:
        try:
            created_by_user = User.objects.get(username='admin')
        except User.DoesNotExist:
            return {'success': False, 'errors': ['Default admin user not found. Please specify created_by_user.'], 'success_count': 0}
    
    # Import Quizzes
    if 'Quizzes' not in wb.sheetnames:
        return {'success': False, 'errors': ['Excel file must contain a "Quizzes" sheet'], 'success_count': 0}
    
    quiz_sheet = wb['Quizzes']
    quizzes_data = {}
    
    for row_idx, row in enumerate(quiz_sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0] or not str(row[0]).strip():  # Skip empty rows
            continue
        
        quiz_title = str(row[0]).strip()
        description = str(row[1]).strip() if row[1] else ''
        
        # Get or create user for created_by
        created_by_username = str(row[2]).strip() if row[2] else 'admin'
        try:
            created_by = User.objects.get(username=created_by_username)
        except User.DoesNotExist:
            created_by = created_by_user
            errors.append(f'Row {row_idx}: User "{created_by_username}" not found, using "{created_by_user.username}"')
        
        is_active_str = str(row[3]).strip().lower() if row[3] else 'yes'
        is_active = is_active_str in ['yes', 'y', 'true', '1']
        
        # Create or get quiz
        quiz, created = Quiz.objects.get_or_create(
            title=quiz_title,
            defaults={
                'description': description,
                'created_by': created_by,
                'is_active': is_active
            }
        )
        if not created:
            quiz.description = description
            quiz.is_active = is_active
            quiz.save()
        
        quizzes_data[quiz_title] = quiz
    
    # Import Questions
    if 'Questions' not in wb.sheetnames:
        errors.append('Excel file should contain a "Questions" sheet')
    else:
        question_sheet = wb['Questions']
        
        for row_idx, row in enumerate(question_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0] or not str(row[0]).strip():  # Skip empty rows
                continue
            
            quiz_title = str(row[0]).strip()
            if quiz_title not in quizzes_data:
                errors.append(f'Questions sheet, Row {row_idx}: Quiz "{quiz_title}" not found in Quizzes sheet')
                continue
            
            quiz = quizzes_data[quiz_title]
            question_text = str(row[1]).strip() if row[1] else ''
            
            if not question_text:
                errors.append(f'Questions sheet, Row {row_idx}: Question text is required')
                continue
            
            question_type_str = str(row[2]).strip() if row[2] else 'Multiple Choice'
            question_type_map = {
                'multiple choice': 'multiple_choice',
                'text answer': 'text',
                'text': 'text',
            }
            question_type = question_type_map.get(question_type_str.lower(), 'multiple_choice')
            
            try:
                order = int(row[3]) if row[3] else 0
            except (ValueError, TypeError):
                order = 0
            
            try:
                time_limit = int(row[4]) if row[4] else 1
            except (ValueError, TypeError):
                time_limit = 1
            
            # Create question
            question, created = Question.objects.get_or_create(
                quiz=quiz,
                question_text=question_text,
                defaults={
                    'question_type': question_type,
                    'order': order,
                    'time_limit_minutes': time_limit
                }
            )
            if not created:
                question.question_type = question_type
                question.order = order
                question.time_limit_minutes = time_limit
                question.save()
    
    # Import Choices
    if 'Choices' in wb.sheetnames:
        choice_sheet = wb['Choices']
        
        for row_idx, row in enumerate(choice_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0] or not str(row[0]).strip():  # Skip empty rows
                continue
            
            quiz_title = str(row[0]).strip()
            if quiz_title not in quizzes_data:
                continue  # Already logged error in Questions import
            
            question_text = str(row[1]).strip() if row[1] else ''
            if not question_text:
                continue
            
            # Find the question
            quiz = quizzes_data[quiz_title]
            try:
                question = Question.objects.get(quiz=quiz, question_text=question_text)
            except Question.DoesNotExist:
                errors.append(f'Choices sheet, Row {row_idx}: Question "{question_text}" not found for quiz "{quiz_title}"')
                continue
            except Question.MultipleObjectsReturned:
                question = Question.objects.filter(quiz=quiz, question_text=question_text).first()
                errors.append(f'Choices sheet, Row {row_idx}: Multiple questions found, using first one')
            
            # Only add choices for multiple choice questions
            if question.question_type != 'multiple_choice':
                continue
            
            choice_text = str(row[2]).strip() if row[2] else ''
            if not choice_text:
                continue
            
            is_correct_str = str(row[3]).strip().lower() if row[3] else 'no'
            is_correct = is_correct_str in ['yes', 'y', 'true', '1']
            
            # Create choice
            Choice.objects.get_or_create(
                question=question,
                choice_text=choice_text,
                defaults={'is_correct': is_correct}
            )
    
    success_count = len(quizzes_data)
    
    return {
        'success': len(errors) == 0,
        'errors': errors,
        'success_count': success_count
    }

